"""
One-time edit: spread the real WMS exports across several famous Indian
fashion brands instead of everything being 'Zudio'.

Every row in the two raw exports (`Inbound Dataset.xlsx`, `Outbound
Dataset.xlsx`) currently carries 'Zudio' (or, for the two BO APPAREL rows,
'ETHNICITY') in every brand column. That's realistic for a single-brand
store but flat for a multi-brand warehouse twin. This script reassigns the
brand columns ('Product Brand', 'Brand', and inbound's 'Store_Master.Brand')
to a curated pool of well-known Indian apparel brands, picked per product
*Style* (not per row) so the same style always keeps the same brand
everywhere it appears -- including the one style ('LETE00186') that shows
up in both inbound and outbound.

Brand choice per style is based on its Division/Family, so the pairing
stays plausible (e.g. kurtas -> Manyavar/FabIndia, western wear ->
Westside/AND/Only, kidswear -> Max Fashion/Gini & Jony), and is picked
deterministically via md5(style) so re-runs are stable. BO APPAREL rows
(already 'ETHNICITY') are left untouched.

Run once, offline: `python scripts/diversify_brands.py`. Edits the two
.xlsx files in place. Run `python scripts/convert_real_data.py` afterwards
to regenerate the app's bundled JSON/CSV from the updated source data.
"""

import hashlib
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

BRAND_COLS = ["Product Brand", "Brand", "Store_Master.Brand"]

MENS_ETHNIC = ["Manyavar", "FabIndia"]
MENS_CASUAL = ["Allen Solly", "Peter England", "Van Heusen", "Zudio"]
WOMENS_WESTERN = ["Westside", "AND", "Only", "Zudio"]
WOMENS_ETHNIC = ["Biba", "W", "FabIndia"]
KIDS = ["Max Fashion", "Gini & Jony", "Zudio Kids"]

ETHNIC_FAMILY_HINTS = ("ETHNIC", "KURTA", "ETHNO")


def pool_for(division: str, family: str) -> list[str] | None:
    division = (division or "").upper()
    family = (family or "").upper()
    if division == "MENS WEAR":
        if any(h in family for h in ETHNIC_FAMILY_HINTS):
            return MENS_ETHNIC
        return MENS_CASUAL
    if division == "WOMENS ETHNICWEAR":
        return WOMENS_ETHNIC
    if division == "WOMENS WESTERNWEAR":
        return WOMENS_WESTERN
    if division == "KIDS WEAR":
        return KIDS
    return None  # BO APPAREL and anything else: leave as-is


def brand_for_style(style: str, division: str, family: str) -> str | None:
    pool = pool_for(division, family)
    if not pool:
        return None
    digest = hashlib.md5(style.encode("utf-8")).hexdigest()
    return pool[int(digest, 16) % len(pool)]


def col_index(header_row, name: str) -> int | None:
    for cell in header_row:
        if cell.value == name:
            return cell.column
    return None


def process(path: Path, style_brand: dict) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    header = next(ws.iter_rows(min_row=1, max_row=1))
    idx = {name: col_index(header, name) for name in ["Style", "Division", "Family", *BRAND_COLS]}
    idx = {k: v for k, v in idx.items() if v is not None}

    for row in ws.iter_rows(min_row=2):
        style_cell = row[idx["Style"] - 1] if "Style" in idx else None
        if style_cell is None or style_cell.value is None:
            continue
        style = str(style_cell.value).strip()
        division = row[idx["Division"] - 1].value if "Division" in idx else None
        family = row[idx["Family"] - 1].value if "Family" in idx else None

        if style not in style_brand:
            style_brand[style] = brand_for_style(style, division, family)
        brand = style_brand[style]
        if brand is None:
            continue  # unmapped division (e.g. BO APPAREL): leave original brand

        for col_name in BRAND_COLS:
            if col_name in idx:
                row[idx[col_name] - 1].value = brand

    wb.save(path)


def main():
    style_brand: dict = {}
    process(ROOT / "Inbound Dataset.xlsx", style_brand)
    process(ROOT / "Outbound Dataset.xlsx", style_brand)

    from collections import Counter

    counts = Counter(b for b in style_brand.values() if b)
    print("Brand mix across styles:")
    for brand, n in counts.most_common():
        print(f"  {brand}: {n} style(s)")
    unmapped = sum(1 for b in style_brand.values() if b is None)
    print(f"  (left unchanged / BO APPAREL etc.): {unmapped} style(s)")


if __name__ == "__main__":
    main()
