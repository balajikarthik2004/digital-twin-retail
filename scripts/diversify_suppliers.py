"""
One-time edit: split the single real inbound goods-in batch across several
real Indian garment-supplier names instead of one vendor for every row.

Before this script, `Inbound Dataset.xlsx` has exactly one ASN
(WMS_ASN_NO == 'GRN-60-Aug26B44') and one vendor
('KB-HYD-WARANGAL HANUMAKONDA') for all 100 rows, so `convert_real_data.py`
always produces exactly one receipt. This script regroups the sheet's 11
distinct product Styles into 5 sub-batches by category, gives each
sub-batch its own synthetic ASN/PO number and a real, well-known Indian
garment-manufacturing/export supplier, and rewrites the identifying
columns (WMS_ASN_NO, VENDOR_ASN_NO, Pono, VendorName, and the WMS_ASN_NO
portion embedded in Asn Qkey) so `build_receipts()` in
convert_real_data.py naturally produces one receipt per supplier.

Run once, offline, AFTER diversify_brands.py:
    python scripts/diversify_suppliers.py
Then regenerate the app's bundled data:
    python scripts/convert_real_data.py
"""

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
ORIGINAL_ASN = "GRN-60-Aug26B44"

# Style -> supplier, grouped by category so each supplier's batch reads like
# a real vendor's specialty (menswear house, womenswear exporter, kidswear
# exporter) rather than a random shuffle.
STYLE_SUPPLIER = {
    "MEL00001": "Arvind Limited",       # MENS WEAR
    "MHT00124": "Arvind Limited",       # MENS WEAR
    "LSK00071": "Gokaldas Exports",     # WOMENS WESTERNWEAR
    "LJ00021": "Gokaldas Exports",      # WOMENS WESTERNWEAR
    "LETE00186": "Shahi Exports",       # WOMENS WESTERNWEAR
    "WACL00002": "Shahi Exports",       # WOMENS WESTERNWEAR
    "JGGT00020": "Pearl Global Industries",  # KIDS WEAR
    "TGTS00109": "Pearl Global Industries",  # KIDS WEAR
    "BGT00018": "Pearl Global Industries",   # KIDS WEAR
    "BCT00003": "KPR Mill",             # KIDS WEAR
    "BCT00010": "KPR Mill",             # KIDS WEAR
}

# One new ASN/PO number per supplier, in the same 'GRN-<n>-Aug26B<seq>' shape
# as the original so it still reads like a WMS-generated code.
SUPPLIER_ASN = {
    "Arvind Limited": "GRN-61-Aug26B01",
    "Gokaldas Exports": "GRN-62-Aug26B02",
    "Shahi Exports": "GRN-63-Aug26B03",
    "Pearl Global Industries": "GRN-64-Aug26B04",
    "KPR Mill": "GRN-65-Aug26B05",
}


def col_index(header_row, name: str) -> int | None:
    for cell in header_row:
        if cell.value == name:
            return cell.column
    return None


def main():
    path = ROOT / "Inbound Dataset.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    header = next(ws.iter_rows(min_row=1, max_row=1))
    idx = {
        name: col_index(header, name)
        for name in ["Style", "WMS_ASN_NO", "VENDOR_ASN_NO", "Pono", "VendorName", "Asn Qkey"]
    }
    missing = [k for k, v in idx.items() if v is None]
    if missing:
        raise SystemExit(f"Expected columns not found: {missing}")

    touched = {}
    for row in ws.iter_rows(min_row=2):
        style = row[idx["Style"] - 1].value
        style = str(style).strip() if style is not None else None
        supplier = STYLE_SUPPLIER.get(style)
        if supplier is None:
            continue  # style not in the mapping: leave original ASN/vendor
        new_asn = SUPPLIER_ASN[supplier]

        row[idx["WMS_ASN_NO"] - 1].value = new_asn
        row[idx["VENDOR_ASN_NO"] - 1].value = new_asn
        row[idx["Pono"] - 1].value = new_asn
        row[idx["VendorName"] - 1].value = supplier

        qkey_cell = row[idx["Asn Qkey"] - 1]
        if qkey_cell.value is not None:
            qkey_cell.value = str(qkey_cell.value).replace(ORIGINAL_ASN, new_asn)

        touched[supplier] = touched.get(supplier, 0) + 1

    wb.save(path)

    print("Supplier batches written:")
    for supplier, asn in SUPPLIER_ASN.items():
        print(f"  {supplier} ({asn}): {touched.get(supplier, 0)} row(s)")


if __name__ == "__main__":
    main()
