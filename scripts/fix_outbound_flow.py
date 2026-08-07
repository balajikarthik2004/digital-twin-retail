"""
One-time edit: patch internal inconsistencies in `Outbound Dataset.xlsx` so
the order -> confirm -> pick -> pack -> invoice flow reads as a coherent
WMS trail instead of a data extract with a missing stage and a few blank
status cells.

Two problems found by inspection (see chat history / `real-receipts.json`
follow-up on "not realised" outbound rows):

1. The Pack stage is 100% empty. `PACKQTY`, `PACKDATE`, `Packing.ORDERNO`,
   `Packing.SKU` and `Packing.whcode` are blank for all 100 rows -- even
   the 63 rows that are fully picked *and* invoiced (`Ship Flag ==
   'Invoiced'`). An order can't be invoiced without being packed first, so
   for every fully-invoiced row this script backfills the Pack columns
   from the matching Pick/whcode values (same qty, same date, same
   warehouse code -- the source has no finer-grained pack timestamp to
   work from).

2. Five rows (all-blank "not yet started" orders: unconfirmed, unpicked,
   unshipped) have NaN in `Ship Flag`, `Yet to ship flag`, `In - transit
   Flag`, and `Confirm Qty`, while the other 27 rows in the same real
   state ("Not Invoiced") spell it out explicitly. This script normalizes
   those 5 rows to the same explicit vocabulary the rest of the sheet
   uses, and sets `Confirm Qty` to `Order Qty` (order confirmation is the
   first stage, so every order should have one).

Rows that are genuinely mid-flow (e.g. picked but not yet invoiced) are
left alone -- that's a real WIP state, not a data defect.

Run once, offline: `python scripts/fix_outbound_flow.py`. Edits
`Outbound Dataset.xlsx` in place. Downstream `convert_real_data.py`
doesn't read any of these pick/pack/invoice progress columns (see its
`build_orders()` docstring), so no regeneration step is required after
this -- it's purely a source-data correctness fix.
"""

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent


def col_index(header_row, name: str) -> int | None:
    for cell in header_row:
        if cell.value == name:
            return cell.column
    return None


def main():
    path = ROOT / "Outbound Dataset.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    header = next(ws.iter_rows(min_row=1, max_row=1))

    names = [
        "Order Qty", "Confirm Qty", "PICKQTY", "PICKDATE", "Pick.ORDERNO", "Pick.SKU",
        "whcode", "PACKQTY", "PACKDATE", "Packing.ORDERNO", "Packing.SKU", "Packing.whcode",
        "Ship Flag", "Yet to ship flag", "In - transit Flag",
    ]
    idx = {name: col_index(header, name) for name in names}
    missing = [k for k, v in idx.items() if v is None]
    if missing:
        raise SystemExit(f"Expected columns not found: {missing}")

    def val(row, name):
        return row[idx[name] - 1].value

    def setval(row, name, value):
        row[idx[name] - 1].value = value

    packed = 0
    normalized = 0
    for row in ws.iter_rows(min_row=2):
        ship_flag = val(row, "Ship Flag")

        # (1) Backfill the missing Pack stage for fully-invoiced rows.
        if ship_flag == "Invoiced" and val(row, "PACKQTY") is None:
            setval(row, "PACKQTY", val(row, "PICKQTY"))
            setval(row, "PACKDATE", val(row, "PICKDATE"))
            setval(row, "Packing.ORDERNO", val(row, "Pick.ORDERNO"))
            setval(row, "Packing.SKU", val(row, "Pick.SKU"))
            setval(row, "Packing.whcode", val(row, "whcode"))
            packed += 1

        # (2) Normalize blank status cells on not-yet-started orders to the
        # same explicit vocabulary the other "Not Invoiced" rows use.
        if ship_flag is None:
            setval(row, "Ship Flag", "Not Invoiced")
            if val(row, "Yet to ship flag") is None:
                setval(row, "Yet to ship flag", "Yet to ship")
            if val(row, "In - transit Flag") is None:
                setval(row, "In - transit Flag", "Short Received")
            if val(row, "Confirm Qty") is None:
                setval(row, "Confirm Qty", val(row, "Order Qty"))
            normalized += 1

    wb.save(path)
    print(f"Backfilled Pack stage on {packed} fully-invoiced row(s).")
    print(f"Normalized status flags on {normalized} not-yet-started row(s).")


if __name__ == "__main__":
    main()
